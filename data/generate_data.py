
"""
Mock Data Generator for Leadership Development Program - 
====================================================================

This script generates a deterministic mock dataset for a statistics tutorial
on causal inference using an observational (self-selection) design.

Design: Open enrollment - voluntary participation with self-selection bias
- Treated:  ~500  managers who voluntarily enrolled (Jan-Mar training)
- Control: ~8,500 managers who did not participate
- Self-selection driven by Organization and performance rating
- No Below / Far Below performers are treated
- Managers grouped into within-organization teams (team_id) for clustered SEs

Author: Generated for Statistics Tutorial
Date: 2026
"""

import numpy as np
import pandas as pd
from datetime import date, timedelta
from scipy import stats
from scipy.special import expit  # logistic function
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# SECTION 1: SETUP AND CONSTANTS
# ============================================================================

SEED = 42
np.random.seed(SEED)

# Sample sizes
N_TOTAL = 9000
N_TREATED_TARGET = 500          # approximate - calibrated via propensity model

# Demographics categories (same as S1)
REGIONS = ['North America', 'Europe', 'Asia Pacific', 'Latin America', 'Middle East & Africa']
ORGANIZATIONS = ['R&D', 'Commercial', 'Manufacturing', 'Digital', 'HR', 'Finance']
JOB_FAMILIES = [
    'Clinical Operations', 'Regulatory Affairs', 'Data Science', 'Medical Affairs',
    'Sales', 'Marketing', 'Supply Chain', 'Quality Assurance', 'Pharmacovigilance',
    'Market Access', 'Human Resources', 'IT & Digital', 'Finance & Accounting',
    'Legal & Compliance', 'Communications'
]
PERFORMANCE_RATINGS = ['Far Below', 'Below', 'Meets', 'Exceeds', 'Far Exceeds']
GENDERS = ['Male', 'Female', 'Non-Binary/Other']

# Survey scale parameters (1-5 Likert)
SURVEY_MIN = 1
SURVEY_MAX = 5

print("=" * 80)
print("LEADERSHIP DEVELOPMENT PROGRAM -  MOCK DATA GENERATOR")
print("=" * 80)
print(f"\nSeed: {SEED}")
print(f"Total Managers: {N_TOTAL} (target ~{N_TREATED_TARGET} treated)")

# ============================================================================
# SECTION 2: GENERATE MANAGER-LEVEL DEMOGRAPHICS
# ============================================================================

print("\n" + "=" * 80)
print("GENERATING MANAGER-LEVEL DATA")
print("=" * 80)

manager_ids = np.arange(1, N_TOTAL + 1)

# Demographics - same distributional assumptions as S1 but for the full population
regions = np.random.choice(REGIONS, size=N_TOTAL)
organizations = np.random.choice(ORGANIZATIONS, size=N_TOTAL)
job_families = np.random.choice(JOB_FAMILIES, size=N_TOTAL)

perf_probs = [0.05, 0.15, 0.60, 0.15, 0.05]
performance = np.random.choice(PERFORMANCE_RATINGS, size=N_TOTAL, p=perf_probs)

gender_probs = [0.48, 0.48, 0.04]
genders = np.random.choice(GENDERS, size=N_TOTAL, p=gender_probs)

ages = np.random.normal(38, 6, N_TOTAL).clip(25, 60).round().astype(int)
tenures = np.random.gamma(3, 4, N_TOTAL).clip(1, 120).round().astype(int)

# num_direct_reports and tot_span_of_control are role attributes that also
# act as HTE modifiers for manager_efficacy (see SECTION 6). They are drawn
# from an auxiliary RNG so that adding / re-ordering them here does NOT
# shift the main seeded stream used for baselines, outcomes, retention, and
# team assignment -- those outputs remain byte-for-byte stable whenever the
# main DGP is unchanged.
_aux_rng = np.random.RandomState(SEED + 1)
num_direct_reports = _aux_rng.randint(5, 13, N_TOTAL)
extra_span = _aux_rng.gamma(2, 5, N_TOTAL).round().astype(int)
tot_span_of_control = np.clip(num_direct_reports + extra_span, num_direct_reports, 50).astype(int)

print(f"[OK] Generated demographics for {N_TOTAL} managers")

# ============================================================================
# SECTION 3: GENERATE PRIOR-YEAR BASELINE SCORES
# ============================================================================

print("\nGenerating prior-year baseline scores...")

def generate_baseline(base_mean, base_sd, n):
    """Generate a baseline score (1-5) as an independent draw; two decimals."""
    scores = np.random.normal(base_mean, base_sd, n)
    return np.clip(np.round(scores, 2), SURVEY_MIN, SURVEY_MAX)

baseline_manager_efficacy = generate_baseline(3.3, 0.85, N_TOTAL)
baseline_workload = generate_baseline(3.0, 0.95, N_TOTAL)
baseline_stay_intention = generate_baseline(2.7, 1.00, N_TOTAL)

print(f"  baseline_manager_efficacy : {baseline_manager_efficacy.mean():.2f}")
print(f"  baseline_workload         : {baseline_workload.mean():.2f}")
print(f"  baseline_stay_intention   : {baseline_stay_intention.mean():.2f}")

# ============================================================================
# SECTION 5: SELF-SELECTION INTO TREATMENT
# ============================================================================

print("\n" + "=" * 80)
print("GENERATING SELF-SELECTION (TREATMENT ASSIGNMENT)")
print("=" * 80)

# Selection drivers: Organization + Performance Rating
# R&D and Digital promoted the programme more heavily - higher participation
org_weight = np.zeros(N_TOTAL)
org_weight[organizations == 'R&D']           =  0.90
org_weight[organizations == 'Digital']       =  0.75
org_weight[organizations == 'Commercial']    =  0.0
org_weight[organizations == 'Manufacturing'] = -0.20
org_weight[organizations == 'HR']            = -0.10
org_weight[organizations == 'Finance']       = -0.25

# Performance rating effect - higher performers are more likely to self-select
# Below and Far Below are hard-blocked from treatment (no treated managers)
perf_weight = np.zeros(N_TOTAL)
perf_weight[performance == 'Far Exceeds'] =  1.20
perf_weight[performance == 'Exceeds']     =  0.60
perf_weight[performance == 'Meets']       =  0.0
perf_weight[performance == 'Below']       = -20.0   # effectively blocks treatment
perf_weight[performance == 'Far Below']   = -20.0   # effectively blocks treatment

# Calibrate intercept so that ~N_TREATED_TARGET managers are treated
# Use bisection for precise calibration
lo, hi = -6.0, 2.0
for _ in range(200):
    intercept = (lo + hi) / 2
    logit_p = intercept + org_weight + perf_weight
    p_treat = expit(logit_p)
    expected_treated = p_treat.sum()
    if expected_treated < N_TREATED_TARGET:
        lo = intercept
    else:
        hi = intercept
    if abs(expected_treated - N_TREATED_TARGET) < 0.5:
        break

print(f"  Calibrated intercept: {intercept:.4f}")
print(f"  Expected treated (sum of probabilities): {p_treat.sum():.1f}")

# Draw treatment from Bernoulli(p_treat)
treatment = (np.random.uniform(size=N_TOTAL) < p_treat).astype(int)

# Hard-enforce: no Below or Far Below treated
treatment[(performance == 'Below') | (performance == 'Far Below')] = 0

n_treated = treatment.sum()
n_control = N_TOTAL - n_treated
print(f"[OK] Treatment assigned: {n_treated} treated, {n_control} control")

# Store propensity scores for diagnostics
propensity_scores = p_treat

# ============================================================================
# SECTION 6: GENERATE MANAGER OUTCOMES
# ============================================================================

print("\n" + "=" * 80)
print("GENERATING MANAGER OUTCOMES")
print("=" * 80)

def generate_outcome_with_baseline(base_mean, base_sd, treatment_effect_d,
                                   treatment, baseline, baseline_r=0.50,
                                   hetero_mask=None, hetero_extra_d=0.0,
                                   hetero_mask_2=None, hetero_extra_d_2=0.0):
    """
    Generate an outcome (1-5) that is correlated with its baseline score
    AND has an independent treatment effect on top.

    Parameters
    ----------
    base_mean       : control-group mean (for units at baseline mean)
    base_sd         : residual SD of outcome
    treatment_effect_d : Cohen's d for the treatment effect
    treatment       : 0/1 treatment indicator
    baseline        : prior-year score (may contain NaN)
    baseline_r      : year-over-year correlation (outcome | baseline)
    hetero_mask     : boolean mask for subgroup with extra effect
    hetero_extra_d  : additional Cohen's d for the subgroup
    hetero_mask_2   : optional second subgroup mask (same scale as hetero_mask)
    hetero_extra_d_2: additional Cohen's d for the second subgroup
    """
    n = len(treatment)

    # Replace NaN baselines with the population mean of non-NaN values
    bl_filled = np.where(np.isnan(baseline), np.nanmean(baseline), baseline)
    bl_centered = bl_filled - np.nanmean(baseline)

    # Outcome = baseline component + treatment effect + noise
    baseline_component = baseline_r * bl_centered
    treatment_effect = treatment_effect_d * base_sd * treatment

    # Heterogeneous treatment effect: hetero_mask may be a 0/1 indicator or a
    # continuous value in [0, 1] that scales the incremental effect; an optional
    # second mask adds a further continuous / binary gradient on top.
    if hetero_mask is not None:
        treatment_effect = treatment_effect + hetero_extra_d * base_sd * treatment * hetero_mask
    if hetero_mask_2 is not None:
        treatment_effect = (
            treatment_effect + hetero_extra_d_2 * base_sd * treatment * hetero_mask_2
        )

    residual_sd = base_sd * np.sqrt(1 - baseline_r ** 2)
    noise = np.random.normal(0, residual_sd, n)

    scores = base_mean + baseline_component + treatment_effect + noise
    scores = np.clip(np.round(scores, 1), SURVEY_MIN, SURVEY_MAX)
    return scores


# Heterogeneous treatment effects on manager_efficacy only (continuous gradients).
# Top HTE:    num_direct_reports (effect grows with span of responsibility).
# Second HTE: tenure_months      (effect is larger for newer managers).
# Both modifiers are scaled to [0, 1] so the incremental Cohen's d adds smoothly
# to the base effect, which gives Causal Forest a continuous signal to split on.
# workload_index_mgr and stay_intention_index_mgr have NO HTE (by design).

# --- Top HTE modifier: num_direct_reports in [5..12] -> scaled in [0, 1]
#     (mean ~0.5). Managers with more direct reports benefit more from training.
nd_reports_scaled = (num_direct_reports.astype(float) - 5.0) / 7.0

# --- Second HTE modifier: tenure_months mapped so that 0 months -> 1.0 and
#     60+ months -> 0.0 (mean ~0.80 because tenure is right-skewed and most
#     managers sit below 60 months). Lower-tenure managers benefit more.
low_tenure_scaled = np.clip(1.0 - tenures.astype(float) / 60.0, 0.0, 1.0)

# --- Manager Efficacy Index: base d 0.33 + continuous HTE gradients.
#     Peak incremental d: +0.15 at max direct reports (12), +0.05 at brand-new
#     tenure (0 months). Average marginal d contribution from HTE ~ 0.075+0.04,
#     so the IPTW marginal Cohen's d stays comparable to the previous design.
manager_efficacy = generate_outcome_with_baseline(
    base_mean=3.4, base_sd=0.90, treatment_effect_d=0.33,
    treatment=treatment, baseline=baseline_manager_efficacy,
    baseline_r=0.60,
    hetero_mask=nd_reports_scaled, hetero_extra_d=0.15,
    hetero_mask_2=low_tenure_scaled, hetero_extra_d_2=0.05,
)
print(
    "  [OK] manager_efficacy_index  (d0.33, +0.15 * direct_reports_scaled [top HTE], "
    "+0.05 * low_tenure_scaled [second HTE])"
)

# --- Workload Index (Manager): no treatment effect (d 0; outcome = baseline + noise)
workload_manager = generate_outcome_with_baseline(
    base_mean=3.2, base_sd=1.00, treatment_effect_d=0.0,
    treatment=treatment, baseline=baseline_workload,
    baseline_r=0.45
)
print("  [OK] workload_index_mgr      (d0, no direct treatment effect)")

# --- Stay Intention Index (Manager): small positive effect (latent d 0.10; IPTW E-value "Weak" 1.5-2.0, not "Very weak")
stay_intention_manager = generate_outcome_with_baseline(
    base_mean=2.8, base_sd=1.00, treatment_effect_d=0.10,
    treatment=treatment, baseline=baseline_stay_intention,
    baseline_r=0.50,
)
print("  [OK] stay_intention_index_mgr (d0.10; IPTW E-value Weak band)")

# ============================================================================
# SECTION 7: GENERATE MANAGER RETENTION OUTCOMES
# ============================================================================

print("\nGenerating manager retention outcomes...")

def generate_retention_with_baseline(base_rate, treatment_or, treatment,
                                     baseline_stay_intent):
    """
    Generate binary retention with treatment OR and a baseline covariate effect.
    Higher baseline_stay_intention (= more intention to stay) -> higher retention.
    """
    base_logit = np.log(base_rate / (1 - base_rate))
    treat_logit = np.log(treatment_or)

    # Centre baseline and add as linear predictor (odds-scale influence)
    bl_centered = baseline_stay_intent - baseline_stay_intent.mean()
    bl_effect = 0.30 * bl_centered  # moderate baseline influence

    logits = base_logit + treat_logit * treatment + bl_effect
    probs = expit(logits)
    return (np.random.uniform(size=len(treatment)) < probs).astype(int)

# Retention DGP is deliberately FRONT-LOADED (non-proportional hazards).
# The treatment effect is concentrated in the first 3 months and attenuates
# to null by 6-9 months. This creates a textbook PH violation (monotone
# decay of the log-hazard-ratio over time) that the Schoenfeld residual
# test should reject and that a time-interaction Cox model recovers cleanly.

# 3-month: base_rate 0.88, OR 4.0 (strong early effect), baseline stay-
# intention coefficient 0.30. Implied period HR ~ 0.27 (highly significant).
retention_3mo = generate_retention_with_baseline(
    base_rate=0.88, treatment_or=4.0, treatment=treatment,
    baseline_stay_intent=baseline_stay_intention,
)

# 6-month: CONDITIONAL on surviving to 3 months.
# Mild residual advantage (treated 0.965 vs control 0.945) -> borderline
# period-specific HR ~0.64 given sparse treated events.
# Target cumulative: treated ~93 %, control ~83 %.
p_surv_6_treated = 0.965
p_surv_6_control = 0.945
cond_prob_6 = np.where(treatment == 1, p_surv_6_treated, p_surv_6_control)
retention_6mo = retention_3mo * (np.random.binomial(1, cond_prob_6, N_TOTAL))

# 9-month: conditional. Treatment effect is NULL from here on.
# Target cumulative: treated ~89 %, control ~79 %.
p_surv_9_treated = 0.955
p_surv_9_control = 0.955
cond_prob_9 = np.where(treatment == 1, p_surv_9_treated, p_surv_9_control)
retention_9mo = retention_6mo * (np.random.binomial(1, cond_prob_9, N_TOTAL))

# 12-month: conditional. Treatment effect is NULL.
# Target cumulative: treated ~85 %, control ~76 %.
p_surv_12_treated = 0.955
p_surv_12_control = 0.955
cond_prob_12 = np.where(treatment == 1, p_surv_12_treated, p_surv_12_control)
retention_12mo = retention_9mo * (np.random.binomial(1, cond_prob_12, N_TOTAL))

# Monotonicity checks
assert (retention_6mo  <= retention_3mo).all(),  "ERROR: non-monotonic 36 mo!"
assert (retention_9mo  <= retention_6mo).all(),  "ERROR: non-monotonic 69 mo!"
assert (retention_12mo <= retention_9mo).all(),  "ERROR: non-monotonic 912 mo!"
print("  [OK] Retention monotonicity verified")

# Internal-only retention flags (not exported to CSV; used for exit_date + descriptives)
_retention_arrays = {
    'retention_3month': retention_3mo,
    'retention_6month': retention_6mo,
    'retention_9month': retention_9mo,
    'retention_12month': retention_12mo,
}

# --- exit_date: yyyy-mm-dd for leavers, blank for retained managers ---
print("\nGenerating exit_date column...")

# Observation window: Jan 2 – Dec 31, 2026  (no one exits on Jan 1)
YEAR_START = date(2026, 1, 2)
YEAR_END   = date(2026, 12, 31)
MAX_DAY    = (YEAR_END - YEAR_START).days  # 363

# Quarter core windows (centre of each exit period)
QUARTER_RANGES = [
    (date(2026, 1, 2),  date(2026, 3, 31)),   # Q1: left before 3-month mark
    (date(2026, 4, 1),  date(2026, 6, 30)),   # Q2: left between 3-6 months
    (date(2026, 7, 1),  date(2026, 9, 30)),   # Q3: left between 6-9 months
    (date(2026, 10, 1), date(2026, 12, 31)),  # Q4: left between 9-12 months
]

# Determine which quarter each leaver falls into
# retention_3month==0                              → exited Q1
# retention_3month==1 & retention_6month==0         → exited Q2
# retention_6month==1 & retention_9month==0         → exited Q3
# retention_9month==1 & retention_12month==0        → exited Q4
# retention_12month==1                              → still employed (blank)
exit_quarter = np.full(N_TOTAL, -1, dtype=int)  # -1 = retained
exit_quarter[retention_3mo == 0] = 0
exit_quarter[(retention_3mo == 1) & (retention_6mo == 0)] = 1
exit_quarter[(retention_6mo == 1) & (retention_9mo == 0)] = 2
exit_quarter[(retention_9mo == 1) & (retention_12mo == 0)] = 3

exit_dates = np.full(N_TOTAL, '', dtype=object)
for q_idx, (q_start, q_end) in enumerate(QUARTER_RANGES):
    mask = exit_quarter == q_idx
    n_leavers = mask.sum()
    if n_leavers == 0:
        continue
    # Draw a uniform base day within the quarter, then add Gaussian jitter
    # (sd ~10 days) so dates bleed slightly across quarter edges for realism
    span_days = (q_end - q_start).days + 1
    base_offsets = np.random.uniform(0, span_days, size=n_leavers)
    jitter = np.random.normal(0, 10, size=n_leavers)
    # Convert to absolute days from YEAR_START; reflect at boundaries
    # to avoid pile-up at Jan 2 / Dec 31
    abs_days = (q_start - YEAR_START).days + base_offsets + jitter
    abs_days = np.where(abs_days < 0, -abs_days, abs_days)
    abs_days = np.where(abs_days > MAX_DAY, 2 * MAX_DAY - abs_days, abs_days)
    abs_days = np.clip(abs_days, 0, MAX_DAY).astype(int)  # safety
    exit_dates[mask] = [(YEAR_START + timedelta(days=int(d))).strftime('%Y-%m-%d')
                        for d in abs_days]

n_leavers_total = (exit_quarter >= 0).sum()
n_retained = (exit_quarter == -1).sum()
print(f"  [OK] exit_date generated: {n_leavers_total} leavers, {n_retained} retained (blank)")

# Survey outcomes (Likert) are not observed for managers who leave by the 6-month mark
# (retention_6mo == 0 covers both Q1 and Q2 leavers). Later leavers still have responses.
survey_missing_6mo = (retention_6mo == 0)
manager_efficacy = manager_efficacy.astype(float)
workload_manager = workload_manager.astype(float)
stay_intention_manager = stay_intention_manager.astype(float)
manager_efficacy[survey_missing_6mo] = np.nan
workload_manager[survey_missing_6mo] = np.nan
stay_intention_manager[survey_missing_6mo] = np.nan
n_blank_survey = int(survey_missing_6mo.sum())
print(f"  [OK] Survey outcomes set to missing for {n_blank_survey} managers (no 6-mo retention)")

# ============================================================================
# SECTION 8: ASSEMBLE MANAGER DATAFRAME
# ============================================================================
# (num_direct_reports and tot_span_of_control are generated in SECTION 2 so
# they can serve as HTE modifiers for manager_efficacy in SECTION 6.)

# --- team_id: group managers into within-organization teams of 5-12 ---
team_ids = np.zeros(N_TOTAL, dtype=int)
team_counter = 1
for org in ORGANIZATIONS:
    org_idx = np.where(organizations == org)[0]
    np.random.shuffle(org_idx)
    i = 0
    while i < len(org_idx):
        remaining = len(org_idx) - i
        if remaining <= 12:
            size = remaining
        else:
            max_size = min(12, remaining - 5)
            size = np.random.randint(5, max_size + 1)
        team_ids[org_idx[i:i + size]] = team_counter
        team_counter += 1
        i += size

n_teams = team_counter - 1
team_sizes_check = pd.Series(team_ids).value_counts()
print(f"\n[OK] Generated {n_teams} teams (within-org, size range: "
      f"{team_sizes_check.min()}-{team_sizes_check.max()})")
print(f"  num_direct_reports range: {num_direct_reports.min()}-{num_direct_reports.max()}, "
      f"mean: {num_direct_reports.mean():.1f}")
print(f"  tot_span_of_control range: {tot_span_of_control.min()}-{tot_span_of_control.max()}, "
      f"mean: {tot_span_of_control.mean():.1f}")
assert (tot_span_of_control >= num_direct_reports).all(), \
    "ERROR: tot_span_of_control < num_direct_reports!"

df_managers = pd.DataFrame({
    'id': manager_ids,
    'team_id': team_ids,
    'treatment': treatment,
    'region': regions,
    'organization': organizations,
    'job_family': job_families,
    'performance_rating': performance,
    'gender': genders,
    'age': ages,
    'tenure_months': tenures,
    'num_direct_reports': num_direct_reports,
    'tot_span_of_control': tot_span_of_control,
    'baseline_manager_efficacy': baseline_manager_efficacy,
    'baseline_workload': baseline_workload,
    'baseline_stay_intention': baseline_stay_intention,
    'propensity_score': np.round(propensity_scores, 4),
    'exit_date': exit_dates,
    'manager_efficacy_index': manager_efficacy,
    'workload_index_mgr': workload_manager,
    'stay_intention_index_mgr': stay_intention_manager,
})

print(f"\n[OK] Manager dataframe assembled: {df_managers.shape}")

# ============================================================================
# SECTION 11: VERIFICATION - SELECTION PATTERNS
# ============================================================================

print("\n" + "=" * 80)
print("VERIFICATION: SELECTION PATTERNS (COVARIATE IMBALANCE)")
print("=" * 80)

# Treatment rates by organisation
print("\nTreatment rate by Organisation:")
org_rates = df_managers.groupby('organization')['treatment'].agg(['sum', 'count', 'mean'])
org_rates.columns = ['n_treated', 'n_total', 'rate']
org_rates = org_rates.sort_values('rate', ascending=False)
for org, row in org_rates.iterrows():
    bar = '-' * int(row['rate'] * 40)
    print(f"  {org:20s}  {row['rate']:.1%}  ({int(row['n_treated']):>3}/{int(row['n_total']):>4})  {bar}")

# Treatment rates by performance rating
print("\nTreatment rate by Performance Rating:")
perf_rates = df_managers.groupby('performance_rating')['treatment'].agg(['sum', 'count', 'mean'])
perf_rates.columns = ['n_treated', 'n_total', 'rate']
perf_rates = perf_rates.reindex(PERFORMANCE_RATINGS)
for rating, row in perf_rates.iterrows():
    bar = '-' * int(row['rate'] * 40)
    print(f"  {rating:20s}  {row['rate']:.1%}  ({int(row['n_treated']):>3}/{int(row['n_total']):>4})  {bar}")

# Baseline means by treatment
print("\nBaseline means by treatment group:")
bl_vars = ['baseline_manager_efficacy', 'baseline_workload',
           'baseline_stay_intention']
for var in bl_vars:
    t_mean = df_managers[df_managers['treatment'] == 1][var].mean()
    c_mean = df_managers[df_managers['treatment'] == 0][var].mean()
    diff = t_mean - c_mean
    print(f"  {var:35s}  Treated: {t_mean:.2f}  Control: {c_mean:.2f}  ={diff:+.2f}")

# Standardised mean differences (SMDs) for all covariates
print("\nStandardised Mean Differences (SMDs):")

def smd_continuous(series_t, series_c):
    """Compute absolute SMD for a continuous variable."""
    pooled_sd = np.sqrt((series_t.std()**2 + series_c.std()**2) / 2)
    if pooled_sd == 0:
        return 0.0
    return (series_t.mean() - series_c.mean()) / pooled_sd

t_df = df_managers[df_managers['treatment'] == 1]
c_df = df_managers[df_managers['treatment'] == 0]

smd_results = []
for var in ['age', 'tenure_months', 'num_direct_reports', 'tot_span_of_control'] + bl_vars:
    s = smd_continuous(t_df[var].dropna(), c_df[var].dropna())
    flag_str = '  *** IMBALANCED' if abs(s) > 0.10 else ''
    print(f"  {var:35s}  SMD = {s:+.3f}{flag_str}")
    smd_results.append({'Variable': var, 'SMD': round(s, 3)})

# Categorical SMDs (difference in proportions / pooled SD)
for var in ['organization', 'region', 'performance_rating', 'gender']:
    dummies = pd.get_dummies(df_managers[var])
    for col in dummies.columns:
        s = smd_continuous(dummies.loc[treatment == 1, col], dummies.loc[treatment == 0, col])
        flag_str = '  *** IMBALANCED' if abs(s) > 0.10 else ''
        print(f"  {var}={col:25s}  SMD = {s:+.3f}{flag_str}")
        smd_results.append({'Variable': f'{var}={col}', 'SMD': round(s, 3)})

# Propensity score distribution
print("\nPropensity Score Distribution:")
for label, mask in [('Treated', treatment == 1), ('Control', treatment == 0)]:
    ps = propensity_scores[mask]
    print(f"  {label:10s}  mean={ps.mean():.3f}  sd={ps.std():.3f}  "
          f"min={ps.min():.3f}  max={ps.max():.3f}  "
          f"Q25={np.percentile(ps, 25):.3f}  Q75={np.percentile(ps, 75):.3f}")

# ============================================================================
# SECTION 12: VERIFICATION - DESCRIPTIVE STATISTICS
# ============================================================================

print("\n" + "=" * 80)
print("DESCRIPTIVE STATISTICS")
print("=" * 80)

print("\n--- MANAGER OUTCOMES ---")
print("\nRetention Rates by Treatment Group:")
for outcome, arr in _retention_arrays.items():
    t_rate = arr[treatment == 1].mean()
    c_rate = arr[treatment == 0].mean()
    print(f"  {outcome}: Treated = {t_rate:.1%}, Control = {c_rate:.1%}, Diff = {(t_rate-c_rate)*100:+.1f} pp")

print("\nSurvey Outcomes by Treatment Group (Mean +/- SD):")
survey_outcomes_mgr = ['manager_efficacy_index', 'workload_index_mgr', 'stay_intention_index_mgr']
for outcome in survey_outcomes_mgr:
    t = df_managers[df_managers['treatment'] == 1][outcome].dropna()
    c = df_managers[df_managers['treatment'] == 0][outcome].dropna()
    print(f"  {outcome}:")
    print(f"    Treated:  {t.mean():.2f}  {t.std():.2f}")
    print(f"    Control:  {c.mean():.2f}  {c.std():.2f}")

# ============================================================================
# SECTION 13: VERIFICATION - STATISTICAL TESTS
# ============================================================================

print("\n" + "=" * 80)
print("STATISTICAL TESTS - CONFIRMING EXPECTED RESULTS")
print("=" * 80)

print("\n--- MANAGER-LEVEL SURVEY OUTCOMES (t-tests) ---")
for outcome in survey_outcomes_mgr:
    t = df_managers[df_managers['treatment'] == 1][outcome].dropna()
    c = df_managers[df_managers['treatment'] == 0][outcome].dropna()
    t_stat, p_val = stats.ttest_ind(t, c)
    d = smd_continuous(t, c)
    sig = '***' if p_val < 0.001 else '**' if p_val < 0.01 else '*' if p_val < 0.05 else 'ns'
    print(f"\n  {outcome}:")
    print(f"    Cohen's d = {d:.3f}")
    print(f"    t = {t_stat:.2f}, p = {p_val:.4f}  {sig}")

print("\n--- MANAGER RETENTION (Logistic Regression) ---")
from sklearn.linear_model import LogisticRegression
for outcome, arr in _retention_arrays.items():
    X = df_managers[['treatment']].values
    y = arr
    model = LogisticRegression(max_iter=1000)
    model.fit(X, y)
    or_val = np.exp(model.coef_[0][0])
    print(f"  {outcome}:  OR = {or_val:.2f}")

# Manager efficacy heterogeneity (ground truth §10)
print("\n--- MANAGER EFFICACY (treated vs control, Cohen's d) ---")
me_t = df_managers[df_managers['treatment'] == 1]['manager_efficacy_index'].dropna()
me_c = df_managers[df_managers['treatment'] == 0]['manager_efficacy_index'].dropna()
d_overall = smd_continuous(me_t, me_c)
print(f"  Overall                    :  d = {d_overall:.3f}  (expected ~0.33-0.45 crude SMD)")

# Top HTE check: managers with large spans (>=9 direct reports) vs small (<=6).
high_span_mask = df_managers['num_direct_reports'] >= 9
low_span_mask  = df_managers['num_direct_reports'] <= 6
hs_t = df_managers[high_span_mask & (df_managers['treatment'] == 1)]['manager_efficacy_index'].dropna()
hs_c = df_managers[high_span_mask & (df_managers['treatment'] == 0)]['manager_efficacy_index'].dropna()
ls_t = df_managers[low_span_mask  & (df_managers['treatment'] == 1)]['manager_efficacy_index'].dropna()
ls_c = df_managers[low_span_mask  & (df_managers['treatment'] == 0)]['manager_efficacy_index'].dropna()
d_high_span = smd_continuous(hs_t, hs_c)
d_low_span  = smd_continuous(ls_t, ls_c)
print(f"  High direct reports (>=9)  :  d = {d_high_span:.3f}  (expected larger than low-span)")
print(f"  Low  direct reports (<=6)  :  d = {d_low_span:.3f}  (expected ~ base 0.33)")

# Second HTE check: low-tenure managers should see a somewhat larger effect.
low_ten_mask  = df_managers['tenure_months'] <= 6
high_ten_mask = df_managers['tenure_months'] >= 24
lt_t = df_managers[low_ten_mask  & (df_managers['treatment'] == 1)]['manager_efficacy_index'].dropna()
lt_c = df_managers[low_ten_mask  & (df_managers['treatment'] == 0)]['manager_efficacy_index'].dropna()
ht_t = df_managers[high_ten_mask & (df_managers['treatment'] == 1)]['manager_efficacy_index'].dropna()
ht_c = df_managers[high_ten_mask & (df_managers['treatment'] == 0)]['manager_efficacy_index'].dropna()
d_low_ten  = smd_continuous(lt_t, lt_c)
d_high_ten = smd_continuous(ht_t, ht_c)
print(f"  Tenure <=6 months          :  d = {d_low_ten:.3f}  (expected slightly larger than high-tenure)")
print(f"  Tenure >=24 months         :  d = {d_high_ten:.3f}  (expected ~ base 0.33)")

print("\n--- STAY INTENTION (treated vs control, Cohen's d) ---")
si_t = df_managers[df_managers['treatment'] == 1]['stay_intention_index_mgr'].dropna()
si_c = df_managers[df_managers['treatment'] == 0]['stay_intention_index_mgr'].dropna()
d_si = smd_continuous(si_t, si_c)
print(f"  stay_intention_index_mgr:  d = {d_si:.3f}  (crude SMD; IPTW marginal d ~0.13–0.16)")

# ============================================================================
# SECTION 14: EXPORT DATA
# ============================================================================

print("\n" + "=" * 80)
print("EXPORTING DATA")
print("=" * 80)

df_managers.to_csv('./manager_data.csv', index=False)

print("\n[OK] Data exported to:")
print("  - ./manager_data.csv")

# ============================================================================
# SECTION 15: EXCEL REPORT
# ============================================================================

print("\n" + "=" * 80)
print("GENERATING EXCEL DESCRIPTIVES REPORT")
print("=" * 80)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Formatting constants (same as S1)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BOLD_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
THIN_BORDER = Border(bottom=Side(style='thin', color='B0B0B0'))

GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def apply_header_format(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def apply_alternating_rows(ws, start_row, end_row, max_col, bold_keywords=None):
    if bold_keywords is None:
        bold_keywords = ['Overall', 'Total']
    for r in range(start_row, end_row + 1):
        fill = ALT_ROW_FILL if (r - start_row) % 2 == 0 else WHITE_FILL
        is_bold = False
        for col in range(1, max_col + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.value is not None and any(kw in str(cell.value) for kw in bold_keywords):
                is_bold = True
        if is_bold:
            for col in range(1, max_col + 1):
                ws.cell(row=r, column=col).font = BOLD_FONT


def auto_fit_columns(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def write_title(ws, row, title_text, max_col=1):
    cell = ws.cell(row=row, column=1, value=title_text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='left')
    if max_col > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)


def write_df(ws, df, start_row, start_col=1):
    for c_idx, col_name in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=c_idx, value=str(col_name))
    apply_header_format(ws, start_row, start_col + len(df.columns) - 1)

    for r_idx, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, value in enumerate(row_data, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx)
            if isinstance(value, (np.integer,)):
                cell.value = int(value)
            elif isinstance(value, (np.floating,)):
                cell.value = round(float(value), 3)
            elif pd.isna(value):
                cell.value = ''
            else:
                cell.value = value

    end_row = start_row + len(df)
    max_col = start_col + len(df.columns) - 1
    apply_alternating_rows(ws, start_row + 1, end_row, max_col)
    return end_row


def apply_pvalue_conditional(ws, col_letter, start_row, end_row):
    for r in range(start_row, end_row + 1):
        cell = ws[f"{col_letter}{r}"]
        try:
            val = float(cell.value)
        except (TypeError, ValueError):
            continue
        if val < 0.05:
            cell.fill = GREEN_FILL
        elif val < 0.10:
            cell.fill = YELLOW_FILL
        else:
            cell.fill = RED_FILL


def descriptives_by_group(df, variables, group_col='treatment',
                          group_labels={1: 'Treated', 0: 'Control'}):
    rows = []
    for var in variables:
        for val, label in list(group_labels.items()) + [('all', 'Overall')]:
            s = df[var].dropna() if val == 'all' else df[df[group_col] == val][var].dropna()
            rows.append({
                'Variable': var, 'Group': label, 'n': len(s),
                'Mean': round(s.mean(), 3), 'SD': round(s.std(), 3),
                'Min': round(s.min(), 3), 'Max': round(s.max(), 3),
                'Median': round(s.median(), 3),
                'Skewness': round(s.skew(), 3),
                'Kurtosis': round(s.kurtosis(), 3),
            })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Create workbook
# ---------------------------------------------------------------------------
EXCEL_PATH = './data_descriptives.xlsx'
wb = Workbook()

# ===== SHEET 1: README =====================================================
ws_readme = wb.active
ws_readme.title = "README"

n_treated_actual = int(treatment.sum())
n_control_actual = int(N_TOTAL - treatment.sum())
readme_lines = [
    ("Leadership Development Program -  Descriptive Report", TITLE_FONT),
    ("", None),
    ("Dataset Overview", BOLD_FONT),
    ("This workbook contains descriptive statistics and selection-pattern diagnostics for an", None),
    ("observational (open-enrollment) evaluation of a leadership development program.", None),
    ("", None),
    ("Design", BOLD_FONT),
    (f" Treated group: {n_treated_actual} managers who voluntarily enrolled (Jan-Mar)", None),
    (f" Control pool : {n_control_actual} managers who did not participate", None),
    (" Self-selection driven by: Organisation, performance rating", None),
    (" Seed for reproducibility: 42", None),
    ("", None),
    (" Prior-year baseline scores included as controls for outcomes", None),
    (" No Below / Far Below performers are treated", None),
    ("", None),
    ("Sheets in This Workbook", BOLD_FONT),
    (" README - This summary sheet", None),
    (" Selection_Patterns - Treatment rates by organisation and performance rating", None),
    (" Covariate_Balance - SMDs for all covariates (pre-weighting)", None),
    (" PS_Summary - Propensity score distribution statistics", None),
    (" Manager_Descriptives - Descriptive stats by treatment group", None),
    (" Retention_Summary - Retention rates at 3, 6, 9, 12 months with chi-square tests", None),
    (" Period_Turnover_Rates - Period-specific turnover rates", None),
    (" Manager_Outcomes - Manager survey outcomes with Cohen's d and t-test p-values", None),
    (" Raw_Managers - Full manager-level dataset", None),
    ("", None),
    ("Formatting Key", BOLD_FONT),
    (" p-value cells: Green (p < .05), Yellow (.05 <= p < .10), Red (p >= .10)", None),
    (" Bold rows indicate 'Overall' or 'Total' summaries", None),
]
for r_idx, (text, font) in enumerate(readme_lines, start=1):
    cell = ws_readme.cell(row=r_idx, column=1, value=text)
    if font:
        cell.font = font
ws_readme.column_dimensions['A'].width = 90
print("  [OK] README")

# ===== SHEET 2: Selection_Patterns ==========================================
ws_sel = wb.create_sheet("Selection_Patterns")
write_title(ws_sel, 1, "Treatment Selection Patterns (Self-Selection Drivers)", max_col=5)

# By Organisation
sel_org_rows = []
for org in sorted(ORGANIZATIONS):
    sub = df_managers[df_managers['organization'] == org]
    n_t = sub['treatment'].sum()
    n_total = len(sub)
    sel_org_rows.append({
        'Organisation': org,
        'n Treated': int(n_t),
        'n Total': int(n_total),
        'Participation Rate %': round(n_t / n_total * 100, 1) if n_total > 0 else 0
    })
sel_org_rows.append({
    'Organisation': 'Overall',
    'n Treated': int(treatment.sum()),
    'n Total': int(N_TOTAL),
    'Participation Rate %': round(treatment.mean() * 100, 1)
})
df_sel_org = pd.DataFrame(sel_org_rows)

row_ptr = 3
write_title(ws_sel, row_ptr, "By Organisation", max_col=4)
row_ptr = write_df(ws_sel, df_sel_org, start_row=row_ptr + 2) + 2

# By Performance Rating
sel_perf_rows = []
for rating in PERFORMANCE_RATINGS:
    sub = df_managers[df_managers['performance_rating'] == rating]
    n_t = sub['treatment'].sum()
    n_total = len(sub)
    sel_perf_rows.append({
        'Performance Rating': rating,
        'n Treated': int(n_t),
        'n Total': int(n_total),
        'Participation Rate %': round(n_t / n_total * 100, 1) if n_total > 0 else 0
    })
df_sel_perf = pd.DataFrame(sel_perf_rows)
write_title(ws_sel, row_ptr, "By Performance Rating", max_col=4)
row_ptr = write_df(ws_sel, df_sel_perf, start_row=row_ptr + 2)

auto_fit_columns(ws_sel)
print("  [OK] Selection_Patterns")

# ===== SHEET 3: Covariate_Balance ===========================================
ws_cov = wb.create_sheet("Covariate_Balance")
write_title(ws_cov, 1, "Covariate Balance  Standardised Mean Differences (Pre-Weighting)", max_col=6)

cov_bal_rows = []
# Continuous
for var in ['age', 'tenure_months', 'num_direct_reports', 'tot_span_of_control',
            'baseline_manager_efficacy', 'baseline_workload',
            'baseline_stay_intention']:
    tv = t_df[var].dropna()
    cv = c_df[var].dropna()
    s = smd_continuous(tv, cv)
    t_stat, p_val = stats.ttest_ind(tv, cv, equal_var=False)
    cov_bal_rows.append({
        'Variable': var,
        'Treated Mean': round(tv.mean(), 3),
        'Control Mean': round(cv.mean(), 3),
        'SMD': round(s, 3),
        't-statistic': round(t_stat, 3),
        'p-value': round(p_val, 4),
    })

# Categorical - proportion difference as SMD
for var in ['organization', 'region', 'performance_rating', 'gender']:
    dummies = pd.get_dummies(df_managers[var], prefix=var)
    for col in dummies.columns:
        tv = dummies.loc[treatment == 1, col].astype(float)
        cv = dummies.loc[treatment == 0, col].astype(float)
        s = smd_continuous(tv, cv)
        t_stat, p_val = stats.ttest_ind(tv, cv, equal_var=False)
        cov_bal_rows.append({
            'Variable': col,
            'Treated Mean': round(tv.mean(), 3),
            'Control Mean': round(cv.mean(), 3),
            'SMD': round(s, 3),
            't-statistic': round(t_stat, 3),
            'p-value': round(p_val, 4),
        })

df_cov_bal = pd.DataFrame(cov_bal_rows)
end_r = write_df(ws_cov, df_cov_bal, start_row=3)
pval_col = get_column_letter(6)
apply_pvalue_conditional(ws_cov, pval_col, 4, end_r)
auto_fit_columns(ws_cov)
print("  [OK] Covariate_Balance")

# ===== SHEET 4: PS_Summary =================================================
ws_ps = wb.create_sheet("PS_Summary")
write_title(ws_ps, 1, "Propensity Score Distribution (Pre-Weighting)", max_col=8)

ps_rows = []
for label, mask in [('Treated', treatment == 1), ('Control', treatment == 0), ('Overall', np.ones(N_TOTAL, dtype=bool))]:
    ps = propensity_scores[mask]
    ps_rows.append({
        'Group': label, 'n': int(mask.sum()),
        'Mean': round(ps.mean(), 4), 'SD': round(ps.std(), 4),
        'Min': round(ps.min(), 4), 'Q25': round(np.percentile(ps, 25), 4),
        'Median': round(np.median(ps), 4), 'Q75': round(np.percentile(ps, 75), 4),
        'Max': round(ps.max(), 4),
    })
df_ps = pd.DataFrame(ps_rows)
end_r = write_df(ws_ps, df_ps, start_row=3)
auto_fit_columns(ws_ps)
print("  [OK] PS_Summary")

# ===== SHEET 5: Manager_Descriptives ========================================
ws_mgr_desc = wb.create_sheet("Manager_Descriptives")
mgr_cont_vars = ['age', 'tenure_months', 'num_direct_reports', 'tot_span_of_control',
                  'baseline_manager_efficacy', 'baseline_workload',
                  'baseline_stay_intention',
                  'manager_efficacy_index', 'workload_index_mgr', 'stay_intention_index_mgr']
df_mgr_desc = descriptives_by_group(df_managers, mgr_cont_vars)
write_title(ws_mgr_desc, 1, "Manager-Level Descriptive Statistics (Continuous & Likert Variables)", max_col=len(df_mgr_desc.columns))
write_df(ws_mgr_desc, df_mgr_desc, start_row=3)
auto_fit_columns(ws_mgr_desc)
print("  [OK] Manager_Descriptives")

# ===== SHEET 6: Retention_Summary ===========================================
ws_ret = wb.create_sheet("Retention_Summary")
write_title(ws_ret, 1, "Retention Rates by Treatment Group with Chi-Square Tests", max_col=10)

ret_rows = []
for outcome, label in [('retention_3month', '3-Month'), ('retention_6month', '6-Month'),
                        ('retention_9month', '9-Month'), ('retention_12month', '12-Month')]:
    arr = _retention_arrays[outcome]
    t_vals = arr[treatment == 1]
    c_vals = arr[treatment == 0]
    ct = pd.crosstab(df_managers['treatment'], pd.Series(arr))
    chi2, p_val, _, _ = stats.chi2_contingency(ct)
    ret_rows.append({
        'Timepoint': label,
        'Treated Retained': int(t_vals.sum()),
        'Treated Lost': int(len(t_vals) - t_vals.sum()),
        'Treated Rate %': round(t_vals.mean() * 100, 1),
        'Control Retained': int(c_vals.sum()),
        'Control Lost': int(len(c_vals) - c_vals.sum()),
        'Control Rate %': round(c_vals.mean() * 100, 1),
        'Difference (pp)': round((t_vals.mean() - c_vals.mean()) * 100, 1),
        'Chi-square': round(chi2, 3),
        'p-value': round(p_val, 4),
    })
df_ret = pd.DataFrame(ret_rows)
end_r = write_df(ws_ret, df_ret, start_row=3)
pval_col = get_column_letter(10)
apply_pvalue_conditional(ws_ret, pval_col, 4, end_r)
auto_fit_columns(ws_ret)
print("  [OK] Retention_Summary")

# ===== SHEET 8: Period_Turnover_Rates ========================================
ws_period = wb.create_sheet("Period_Turnover_Rates")
write_title(ws_period, 1, "Period-Specific Turnover Rates Analysis", max_col=10)

t_mgr = df_managers[df_managers['treatment'] == 1]
c_mgr = df_managers[df_managers['treatment'] == 0]
n_t = len(t_mgr)
n_c = len(c_mgr)

period_data = []
prev_t_retained, prev_c_retained = n_t, n_c
for i, (col, lbl) in enumerate([
    ('retention_3month', '0-3 Months'), ('retention_6month', '3-6 Months'),
    ('retention_9month', '6-9 Months'),  ('retention_12month', '9-12 Months')]):

    arr = _retention_arrays[col]
    curr_t = int(arr[treatment == 1].sum())
    curr_c = int(arr[treatment == 0].sum())
    lost_t = prev_t_retained - curr_t
    lost_c = prev_c_retained - curr_c
    rate_t = lost_t / prev_t_retained * 100 if prev_t_retained > 0 else 0
    rate_c = lost_c / prev_c_retained * 100 if prev_c_retained > 0 else 0

    contingency = [[curr_t, lost_t], [curr_c, lost_c]]
    try:
        chi2, p_val, _, _ = stats.chi2_contingency(contingency)
    except ValueError:
        chi2, p_val = 0, 1.0

    period_data.append({
        'Period': lbl,
        'Treated At Risk': prev_t_retained, 'Treated Lost': lost_t,
        'Treated Turnover %': round(rate_t, 1),
        'Control At Risk': prev_c_retained, 'Control Lost': lost_c,
        'Control Turnover %': round(rate_c, 1),
        'Difference (pp)': round(rate_c - rate_t, 1),
        'Chi-square': round(chi2, 3), 'p-value': round(p_val, 4)
    })
    prev_t_retained = curr_t
    prev_c_retained = curr_c

df_period = pd.DataFrame(period_data)
end_r = write_df(ws_period, df_period, start_row=3)
pval_col = get_column_letter(10)
apply_pvalue_conditional(ws_period, pval_col, 4, end_r)
auto_fit_columns(ws_period)
print("  [OK] Period_Turnover_Rates")

# ===== SHEET 9: Manager_Outcomes ============================================
ws_mgr_out = wb.create_sheet("Manager_Outcomes")
write_title(ws_mgr_out, 1, "Manager-Level Survey Outcomes  Treatment vs Control", max_col=10)

mgr_out_rows = []
for outcome in survey_outcomes_mgr:
    t_vals = df_managers[df_managers['treatment'] == 1][outcome].dropna()
    c_vals = df_managers[df_managers['treatment'] == 0][outcome].dropna()
    t_stat, p_val = stats.ttest_ind(t_vals, c_vals)
    d = smd_continuous(t_vals, c_vals)
    mgr_out_rows.append({
        'Outcome': outcome,
        'Treated n': len(t_vals), 'Treated Mean': round(t_vals.mean(), 3),
        'Treated SD': round(t_vals.std(), 3),
        'Control n': len(c_vals), 'Control Mean': round(c_vals.mean(), 3),
        'Control SD': round(c_vals.std(), 3),
        "Cohen's d": round(d, 3),
        't-statistic': round(t_stat, 3), 'p-value': round(p_val, 4),
    })
df_mgr_out = pd.DataFrame(mgr_out_rows)
end_r = write_df(ws_mgr_out, df_mgr_out, start_row=3)
pval_col = get_column_letter(10)
apply_pvalue_conditional(ws_mgr_out, pval_col, 4, end_r)
auto_fit_columns(ws_mgr_out)
print("  [OK] Manager_Outcomes")

# ===== SHEET 8: Raw_Managers ===============================================
ws_raw_mgr = wb.create_sheet("Raw_Managers")
write_title(ws_raw_mgr, 1, "Full Manager-Level Dataset", max_col=len(df_managers.columns))
write_df(ws_raw_mgr, df_managers, start_row=3)
auto_fit_columns(ws_raw_mgr)
print("  [OK] Raw_Managers")

# ---------------------------------------------------------------------------
# Save workbook
# ---------------------------------------------------------------------------
wb.save(EXCEL_PATH)
print(f"\n[DONE] Excel report saved to: {EXCEL_PATH}")

print("\n" + "=" * 80)
print("DATA GENERATION COMPLETE")
print("=" * 80)
print(f"\nFiles created:")
print(f"  ./manager_data.csv        ({df_managers.shape[0]} rows x {df_managers.shape[1]} cols)")
print(f"  ./data_descriptives.xlsx  (8 sheets)")

